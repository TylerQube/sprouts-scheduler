from openpyxl import Workbook
from openpyxl.styles import PatternFill
import math


def build_spreadsheet(shifts):
    wb = Workbook()
    ws = wb.active

    cafe_shifts = list(filter(lambda s: s.initiative == "SPROUTS CAFE HELPER", shifts))
    next_row = fill_cafe(cafe_shifts, wb)

    prep_shifts = list(filter(lambda s: s.initiative == "PREP", shifts))
    next_row = fill_prep(next_row, prep_shifts, wb) + 2

    ce_shifts = list(filter(lambda s: s.initiative == "COMMUNITY EATS SERVER", shifts))
    fill_community_eats(next_row, 1, ce_shifts, wb)

    pm_shifts = list(filter(lambda s: s.initiative == "PRODUCE POSSE" or s.initiative == "SET-UP SQUAD", shifts))
    fill_produce_market(next_row, 7, pm_shifts, wb)

    sm_shifts = list(filter(lambda s: s.initiative == "DONATION DRIVER", shifts))
    next_row = fill_sproutsmobile(next_row, 4, sm_shifts, wb)

    fridge_shifts = list(filter(lambda s: s.initiative == "STOCKING SQUAD" or s.initiative == "CLEANUP CREW", shifts))
    fill_fridge(next_row, 4, fridge_shifts, wb)


    wb.save("output.xlsx")


def fill_cell(cell, color):
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

f_colors = ["FFFF59", "FFFFA6", "FFFF83"]
def fill_fridge(start_row, start_col, f_shifts, wb):
    ws = wb.active
    row = start_row
    col = start_col 
    header = ws.cell(row, col)
    header.value = "COMMUNITY FRIDGE"
    for i in range(col, col + 2):
        fill_cell(ws.cell(row, i), f_colors[0])
    row += 1

    # fill in shifts
    ci = 2
    time_shifts = [s for s in f_shifts if s.day != "n/a"]
    time_shifts.sort(key=lambda s: (
        ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"].index(s.day),
        ["STOCKING SQUAD", "CLEANUP CREW"].index(s.initiative)
    ))
    for i in range(len(f_shifts)):
        s = f_shifts[i]
        c = ws.cell(row, col+1)
        c.value = f"{s.day + " " if s.day != "n/a" else ""}{s.initiative}"
        fill_cell(c, f_colors[ci])
        c = ws.cell(row, col)
        fill_cell(c, f_colors[ci])
        row += 1
        c = ws.cell(row, col)
        c.value = f"{s.time}"
        for iv, v in enumerate(s.volunteers):
            c = ws.cell(row + iv, col+1)
            c.value = v
            fill_cell(c, f_colors[ci])
            fill_cell(ws.cell(row + iv, col), f_colors[ci])
        row += len(s.volunteers)
        ci += 1
        if ci == 3:
            ci = 1

pm_colors = ["FFB17B", "FFE1CB", "FFC194"]
def fill_produce_market(start_row, start_col, pm_shifts, wb):
    ws = wb.active
    row = start_row
    col = start_col 
    header = ws.cell(row, col)
    header.value = "PRODUCE MARKET"
    for i in range(col, col + 2):
        fill_cell(ws.cell(row, i), pm_colors[0])
    row += 1

    # fill in shifts
    ci = 1
    pm_shifts.sort(key=lambda s: (
        ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"].index(s.day),
        ["SET-UP SQUAD", "PRODUCE POSSE"].index(s.initiative)
    ))
    cur_day = ""
    for i in range(len(pm_shifts)):
        s = pm_shifts[i]
        if cur_day != s.day:
            ci += 1
            if ci == 3:
                ci = 1
            # draw day header
            c = ws.cell(row, col)
            fill_cell(c, pm_colors[ci])
            c = ws.cell(row, col+1)
            fill_cell(c, pm_colors[ci])
            c.value = s.day
            row += 1
        cur_day = s.day
        c = ws.cell(row, col)
        c.value = f"{s.time} {s.initiative}"
        for iv, v in enumerate(s.volunteers):
            c = ws.cell(row + iv, col+1)
            c.value = v
            fill_cell(c, pm_colors[ci])
            fill_cell(ws.cell(row + iv, col), pm_colors[ci])
        row += len(s.volunteers)

    return start_row

sm_colors = ["FF7BAC", "FFA6C7", "FFC9DE"]
def fill_sproutsmobile(start_row, start_col, sm_shifts, wb):
    ws = wb.active
    row = start_row
    col = start_col 
    header = ws.cell(row, col)
    header.value = "SPROUTSMOBILE"
    for i in range(col, col + 2):
        fill_cell(ws.cell(row, i), sm_colors[0])
    row += 1

    # fill header
    fill_cell(ws.cell(row, col), sm_colors[1])
    c = ws.cell(row, col+1)
    fill_cell(c, sm_colors[1])
    c.value = "Wednesday"
    row += 1

    # fill in shifts
    ci = 2
    time_shifts = [s for s in sm_shifts if s.day.lower() != "n/a"]
    for i in range(len(time_shifts)):
        s = time_shifts[i]
        ws.cell(row, col).value = s.time
        for iv in range(s.capacity):
            c = ws.cell(row + iv, col+1)
            if iv < len(s.volunteers):
                c.value = s.volunteers[iv]
            fill_cell(c, sm_colors[ci])
            fill_cell(ws.cell(row + iv, col), sm_colors[ci])
        row += s.capacity
        ci += 1
        if ci == 3:
            ci = 1
    
    # fill on-call header
    fill_cell(ws.cell(row, col), sm_colors[1])
    c = ws.cell(row, col+1)
    fill_cell(c, sm_colors[1])
    c.value = "ON CALL DRIVERS"
    row += 1
    c = ws.cell(row, col)
    fill_cell(c, sm_colors[1])
    c.value = "ON CALL DRIVERS"
    on_call = [s for s in sm_shifts if s.day.lower() == "n/a"][0]
    for iv, v in enumerate(on_call.volunteers):
        c = ws.cell(row + iv, col+1)
        c.value = v
        fill_cell(c, sm_colors[ci])
        fill_cell(ws.cell(row + iv, col), sm_colors[ci])
    row += len(on_call.volunteers)

    return row + 1


ce_colors = ["63E5FF", "9BEEFF", "C6F6FF"]
def fill_community_eats(start_row, start_col, ce_shifts, wb):
    ws = wb.active
    row = start_row
    col = start_col 
    header = ws.cell(row, col)
    header.value = "SPROUTS COMMUNITY EATS"
    for i in range(col, col + 2):
        fill_cell(ws.cell(row, i), ce_colors[0])
    row += 1

    # fill header
    fill_cell(ws.cell(row, col), ce_colors[1])
    c = ws.cell(row, col+1)
    fill_cell(c, ce_colors[1])
    c.value = "Friday"
    row += 1

    # fill in shifts
    ci = 2
    ce_shifts.sort(key=lambda s: (["9am-11am", "11am-1pm", "1pm-3:30pm"].index(s.time)))
    for i in range(len(ce_shifts)):
        s = ce_shifts[i]
        ws.cell(row, col).value = s.time
        for iv, v in enumerate(s.volunteers):
            c = ws.cell(row + iv, col+1)
            c.value = v
            fill_cell(c, ce_colors[ci])
            fill_cell(ws.cell(row + iv, col), ce_colors[ci])
        row += len(s.volunteers)
        ci += 1
        if ci == 3:
            ci = 1
    return start_row


prep_colors = ["D684FF", "E2A9FF", "EBC3FF"]
def fill_prep(start_row, prep_shifts, wb):
    ws = wb.active
    row = start_row
    col = 1
    header = ws.cell(row, col)
    header.value = "PREP"
    for i in range(6):
        fill_cell(ws.cell(row, i + 1), prep_colors[0])
    row += 1

    # sort in grid order
    prep_shifts.sort(key=lambda s: (["Monday", "Tuesday", "Wednesday", "Thursday", "Sunday"].index(s.day)))

    # fill header
    color = prep_colors[1]
    fill_cell(ws.cell(row, 1), color)
    for i in range(5):
        c = ws.cell(row, i + 2)
        fill_cell(c, color)
        c.value = ["Monday", "Tuesday", "Wednesday", "Thursday", "Sunday"][i]
    row += 1
    ws.cell(row, 1).value = "5pm-8pm"

    # fill in shifts
    color = prep_colors[2]
    for i in range(len(prep_shifts)):
        s = prep_shifts[i]
        ws.cell(row, col).value = s.time
        for iv, v in enumerate(s.volunteers):
            c = ws.cell(row + iv, i + 2)
            c.value = v
            fill_cell(ws.cell(row + iv, i + 1), color)
            fill_cell(c, color)
    max_volunteers_in_row = len(
        max(prep_shifts, key=lambda s: len(s.volunteers)).volunteers
    )
    row += prep_shifts[0].capacity
    return row


# fill sheet with cafe helper shift assignments
# returns - row at bottom of filled shifts
def fill_cafe(cafe_shifts, wb):
    ws = wb.active
    row = 1
    col = 1
    header = ws.cell(row, col)
    header.value = "SPROUTS CAFE DAYTIME SHIFTS"

    # sort in grid order
    cafe_shifts.sort(
        key=lambda s: (
            ["9am-11am", "11am-1pm", "1pm-3pm", "3pm-5pm"].index(s.time),
            ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"].index(s.day),
        )
    )

    for i in range(4):
        ws.cell(2, (i + 1) * 2).value = ["Monday", "Tuesday", "Wednesday", "Thursday"][
            i
        ]
    for i in range(8):
        fill_cell(ws.cell(2, i + 1), "BBFFAC")
        fill_cell(ws.cell(1, i + 1), "9BFF84")

    # widen volunteer columns
    for l in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[l].width *= 1.5

    # fill in shifts
    colors = ["D8FFD0", "BBFFAC"]
    ci = 0
    row = 3
    for i in range(len(cafe_shifts) + 1):
        if i > 0 and i % 4 == 0 or i >= len(cafe_shifts):
            row += max(cafe_shifts, key=lambda s: s.capacity).capacity
            ci = (ci + 1) % 2
            if i >= len(cafe_shifts):
                break

        row_shifts = cafe_shifts[math.floor(i / 4) : math.floor(i / 4) + 4]

        s = cafe_shifts[i]
        ws.cell(row, col).value = s.time
        for iv in range(8):
            c = ws.cell(row + iv, col + 1)
            if iv < len(s.volunteers):
                c.value = s.volunteers[iv]
            fill_cell(c, colors[ci])
            c = ws.cell(row + iv, col)
            fill_cell(c, colors[ci])

        col += 2
        col = (col - 1) % 8 + 1
    return row + 2
